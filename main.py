"""
贝壳找房周边小区人口估算爬虫（支持任意城市）
用法：
  python main.py          GUI 模式（弹出窗口交互）
  python main.py --cli    命令行模式（原有分步交互式）
  python main.py --test   测试模式，仅抓取1个小区并打印结果
  python main.py --reset  清除所有断点文件，从头开始
"""

import json
import re
import sys
import logging
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.comments import Comment

import config
import runtime_state
from scraper import (
    close_driver,
    get_community_urls,
    get_community_detail,
    sync_browser_cookies,
    set_callbacks as _set_scraper_callbacks,
)
from distance import haversine
from calculator import calc_avg_rooms, calc_population

logging.basicConfig(level=logging.INFO, format="%(levelname)s  %(message)s")
logger = logging.getLogger(__name__)

CHECKPOINT_DIR = Path("checkpoints")
STEP1_FILE = CHECKPOINT_DIR / "step1_urls.json"
STEP2_FILE = CHECKPOINT_DIR / "step2_details.json"


# ──────────────────────────────────────────
# 工具函数
# ──────────────────────────────────────────

def set_checkpoint_dir(path) -> None:
    global CHECKPOINT_DIR, STEP1_FILE, STEP2_FILE
    CHECKPOINT_DIR = Path(path).expanduser()
    STEP1_FILE = CHECKPOINT_DIR / "step1_urls.json"
    STEP2_FILE = CHECKPOINT_DIR / "step2_details.json"


def get_checkpoint_dir() -> Path:
    return CHECKPOINT_DIR


def get_checkpoint_files() -> tuple[Path, Path]:
    return STEP1_FILE, STEP2_FILE


def _save_json(path: Path, data):
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def _load_json(path: Path):
    return json.loads(path.read_text(encoding="utf-8"))


def _prompt_cli_data_dir(log_fn=print) -> str:
    default_dir = runtime_state.get_saved_data_dir(str(CHECKPOINT_DIR))
    print(f"请输入数据保存文件夹 [直接回车使用 {default_dir}]: ", end="", flush=True)
    selected = input().strip() or default_dir
    set_checkpoint_dir(selected)
    log_fn(f"数据保存文件夹: {get_checkpoint_dir()}")
    return str(get_checkpoint_dir())


def _ask_continue(next_step_desc: str, preview_lines: list = None,
                  log_fn=print, ask_fn=None) -> None:
    """
    询问是否继续下一步。
    GUI 模式：ask_fn 为 ScraperGUI.ask_continue，弹窗交互。
    CLI 模式：ask_fn=None，回落到 input() 交互。
    """
    if ask_fn is not None:
        choice = ask_fn(next_step_desc, preview_lines)
        if choice == "n":
            log_fn("已保存断点文件，下次运行将自动跳过已完成的步骤。")
            sys.exit(0)
        if choice == "q":
            sys.exit(0)
        return

    # ── CLI 模式 ──
    if preview_lines:
        log_fn("  预览（前3条）：")
        for line in preview_lines[:3]:
            log_fn(f"    {line}")
    print(f"\n→ 继续 {next_step_desc}？ [y=继续 / n=保存并退出 / q=退出]: ", end="", flush=True)
    while True:
        choice = input().strip().lower()
        if choice == "y":
            return
        if choice == "n":
            log_fn("已保存断点文件，下次运行将自动跳过已完成的步骤。")
            sys.exit(0)
        if choice == "q":
            sys.exit(0)
        print("  请输入 y / n / q: ", end="", flush=True)


# ──────────────────────────────────────────
# Step 1：获取小区 URL 列表
# ──────────────────────────────────────────

def run_step1(log_fn=print, ask_fn=None) -> list:
    if STEP1_FILE.exists():
        data = _load_json(STEP1_FILE)
        urls = data["urls"]
        log_fn(f"\n✅ 从断点恢复 Step 1：共 {len(urls)} 个小区 URL（跳过爬取）")
        return urls

    log_fn("\n━━━ Step 1：搜索贝壳小区列表 ━━━")
    log_fn(f"搜索行政区: {', '.join(config.SEARCH_DISTRICTS)}")

    all_communities = []
    seen_urls = set()
    for district in config.SEARCH_DISTRICTS:
        log_fn(f"\n▶ 行政区: {district}")
        comms = get_community_urls(district)
        for c in comms:
            if c["url"] not in seen_urls:
                seen_urls.add(c["url"])
                all_communities.append(c)
        log_fn(f"  共找到 {len(comms)} 个小区")

    total = len(all_communities)

    with_coords = [c for c in all_communities if c["lng"] is not None]
    without_coords = [c for c in all_communities if c["lng"] is None]

    if with_coords:
        in_range_urls = []
        skipped = 0
        for c in with_coords:
            dist = haversine(config.TARGET_LNG, config.TARGET_LAT, c["lng"], c["lat"])
            if dist <= config.RADIUS_KM:
                in_range_urls.append(c["url"])
            else:
                skipped += 1
        candidate_urls = in_range_urls + [c["url"] for c in without_coords]
        log_fn(f"\n预筛选（列表页坐标）：已排除 {skipped} 个超出 {config.RADIUS_KM}km 的小区")
        log_fn(f"  Step 2 需抓取详情：{len(candidate_urls)} 个（原 {total} 个，节省 {total - len(candidate_urls)} 次请求）")
    else:
        candidate_urls = [c["url"] for c in all_communities]
        log_fn(f"\n列表页未获取到坐标，Step 2 将抓取全部 {len(candidate_urls)} 个详情页")

    _save_json(STEP1_FILE, {"urls": candidate_urls})

    log_fn(f"\n[Step 1 完成] 共找到 {total} 个小区，{len(candidate_urls)} 个进入 Step 2")
    preview = [f"├ {u}" for u in candidate_urls]
    _ask_continue("Step 2（逐一抓取详情 + 距离筛选）", preview, log_fn=log_fn, ask_fn=ask_fn)
    return candidate_urls


# ──────────────────────────────────────────
# Step 2：抓取详情 + Haversine 距离筛选
# ──────────────────────────────────────────

def run_step2(urls: list, log_fn=print, ask_fn=None, max_communities=None) -> list:
    if STEP2_FILE.exists():
        data = _load_json(STEP2_FILE)
        communities = data["communities"]
        log_fn(f"\n✅ 从断点恢复 Step 2：共 {len(communities)} 个小区数据已保存（跳过爬取）")
        return communities

    log_fn(f"\n━━━ Step 2：抓取小区详情页 ━━━\n")

    if max_communities is not None and max_communities < len(urls):
        log_fn(f"⚙ 已设置爬取上限：仅处理前 {max_communities} 个小区（共 {len(urls)} 个）")
        urls = urls[:max_communities]

    all_communities = []
    total = len(urls)

    for i, url in enumerate(urls, 1):
        detail = get_community_detail(url)
        if detail is None:
            log_fn(f"  [{i:>4}/{total}] 跳过（数据获取失败）")
            continue
        all_communities.append(detail)
        log_fn(f"  [{i:>4}/{total}] {detail['name']}  N={detail['N']}  a={detail['a']}")

    _save_json(STEP2_FILE, {"communities": all_communities})

    log_fn(f"\n[Step 2 完成] 扫描 {total} 个 URL，成功获取 {len(all_communities)} 个小区数据")
    _ask_continue("Step 3（计算距离 + 筛选 + 输出 Excel）", log_fn=log_fn, ask_fn=ask_fn)
    return all_communities


# ──────────────────────────────────────────
# Step 3：计算人口 + 输出 Excel
# ──────────────────────────────────────────

def run_step3(communities: list, log_fn=print) -> None:
    safe_name = re.sub(r'[\\/:*?"<>|]', '_', config.TARGET_NAME)
    output_file = f"{safe_name}_人口估算结果.xlsx"
    log_fn(f"\n━━━ Step 3：计算距离 + 筛选 + 输出 {output_file} ━━━")
    log_fn(f"目标坐标: ({config.TARGET_LNG}, {config.TARGET_LAT})  [{config.TARGET_NAME}]")

    for c in communities:
        dist = haversine(config.TARGET_LNG, config.TARGET_LAT, c["lng"], c["lat"])
        c["distance_km"] = round(dist, 3)

    in_range = [c for c in communities if c["distance_km"] <= config.RADIUS_KM]
    log_fn(f"共 {len(communities)} 个小区，{config.RADIUS_KM}km 内 {len(in_range)} 个参与计算\n")

    rows = []
    total_p = 0

    for c in sorted(communities, key=lambda x: x["distance_km"]):
        is_in_range = c["distance_km"] <= config.RADIUS_KM
        b_raw = calc_avg_rooms(c.get("room_dist") or {})
        if b_raw is not None:
            b = round(b_raw, 2)
            b_note = None
        else:
            b = config.DEFAULT_B
            b_note = f"页面无户型数据，使用默认值 {config.DEFAULT_B}"

        N = c.get("N") or 0
        a = c.get("a") or 0

        if is_in_range:
            P = round(calc_population(b, a))
            total_p += P
            b_flag = f"  ⚠ {b_note}" if b_note else ""
            log_fn(f"  ✓ {c['name']}: a={a}, b={b} → P={P:,}{b_flag}")
        else:
            P = None

        rows.append({
            "name": c["name"],
            "distance_km": c["distance_km"],
            "in_range": is_in_range,
            "N": N,
            "a": a,
            "b": b,
            "b_note": b_note,
            "P": P,
            "url": c["url"],
        })

    _write_excel(rows, total_p, output_file)

    log_fn(f"\n[Step 3 完成] 已生成 {output_file}")
    log_fn(f"  总小区数: {len(rows)}（含 {len(rows) - len(in_range)} 个超出范围，灰显）")
    log_fn(f"  {config.RADIUS_KM}km 内: {len(in_range)} 个")
    log_fn(f"  P总 = {total_p:,} 人")


def _write_excel(rows: list, total_p: int, output_file: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{config.CITY_NAME}人口估算"[:31]

    headers = ["小区名称", "距离(km)", f"≤{config.RADIUS_KM}km", "楼栋数N", "总户数a", "平均室数b", "估算人口P", "贝壳链接"]
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align

    ws.row_dimensions[1].height = 20

    gray_fill = PatternFill(start_color="EBEBEB", end_color="EBEBEB", fill_type="solid")
    gray_font = Font(color="888888")

    for row_idx, r in enumerate(rows, 2):
        in_range = r["in_range"]

        def _cell(col, value, _in_range=in_range):
            c = ws.cell(row=row_idx, column=col, value=value)
            if not _in_range:
                c.fill = gray_fill
                c.font = gray_font
            return c

        _cell(1, r["name"])
        _cell(2, r["distance_km"])
        _cell(3, "✓" if in_range else "✗")
        _cell(4, r["N"])
        _cell(5, r["a"])

        b_cell = _cell(6, r["b"])
        if r["b_note"] and in_range:
            b_cell.comment = Comment(r["b_note"], "爬虫")
            b_cell.font = Font(color="AA0000")

        _cell(7, r["P"])

        link_cell = ws.cell(row=row_idx, column=8, value=r["url"])
        link_cell.hyperlink = r["url"]
        if in_range:
            link_cell.font = Font(color="0563C1", underline="single")
        else:
            link_cell.fill = gray_fill
            link_cell.font = Font(color="AAAAAA", underline="single")

    total_row = len(rows) + 2
    summary_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    total_label = ws.cell(row=total_row, column=1, value="合  计（范围内）")
    total_label.font = Font(bold=True)
    total_label.fill = summary_fill

    in_range_count = sum(1 for r in rows if r["in_range"])
    ws.cell(row=total_row, column=3, value=f"{in_range_count} 个").fill = summary_fill

    for col in range(2, 8):
        ws.cell(row=total_row, column=col).fill = summary_fill

    p_total_cell = ws.cell(row=total_row, column=7, value=total_p)
    p_total_cell.font = Font(bold=True)
    p_total_cell.fill = summary_fill

    col_widths = {"A": 28, "B": 10, "C": 10, "D": 10, "E": 10, "F": 12, "G": 12, "H": 55}
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    ws.freeze_panes = "A2"
    wb.save(output_file)


# ──────────────────────────────────────────
# 入口
# ──────────────────────────────────────────

def _main_impl(log_fn=print, ask_continue_fn=None, ask_checkpoint_fn=None,
               verify_fn=None, max_communities=None, data_dir=None):
    """
    爬虫主流程。
    GUI 模式：log_fn / ask_continue_fn / ask_checkpoint_fn / verify_fn 由 ScraperGUI 注入。
    CLI 模式：保持默认值，回落到 print/input 交互。
    """
    _set_scraper_callbacks(log_fn=log_fn, verify_fn=verify_fn)
    args = sys.argv[1:]

    saved_state = runtime_state.load_state()
    if saved_state:
        runtime_state.apply_state_to_config(config, saved_state)

    if data_dir is not None:
        set_checkpoint_dir(data_dir)
    elif "--test" not in args:
        data_dir = _prompt_cli_data_dir(log_fn=log_fn)

    if "--test" not in args:
        runtime_state.save_state(
            runtime_state.state_from_config(config, str(get_checkpoint_dir()), max_communities)
        )

    # ── 测试模式 ──
    if "--test" in args:
        log_fn("🔧 测试模式：只处理第一个找到的小区\n")
        for district in config.SEARCH_DISTRICTS:
            urls = get_community_urls(district)
            if not urls:
                continue
            detail = get_community_detail(urls[0])
            if detail:
                dist = haversine(
                    config.TARGET_LNG, config.TARGET_LAT,
                    detail["lng"], detail["lat"]
                )
                detail["distance_km"] = round(dist, 3)
                log_fn("\n── 抓取结果 ──")
                log_fn(json.dumps(detail, ensure_ascii=False, indent=2))
            break
        return

    # ── 重置断点 ──
    if "--reset" in args:
        for f in get_checkpoint_files():
            if f.exists():
                f.unlink()
                log_fn(f"已删除断点文件: {f}")
        log_fn("断点已清除，重新运行 python main.py 从头开始。")
        return

    # ── 正常流程 ──
    log_fn(f"=== {config.CITY_NAME}周边小区人口估算爬虫 ===")
    log_fn(f"目标地点: {config.TARGET_NAME}")
    log_fn(f"搜索半径: {config.RADIUS_KM} km")
    log_fn(f"目标坐标: ({config.TARGET_LNG}, {config.TARGET_LAT})")
    log_fn(f"数据保存文件夹: {get_checkpoint_dir()}")
    log_fn("坐标说明：小区坐标来自贝壳页面，程序直接用 Haversine 计算距离；目标坐标请尽量与贝壳坐标保持同一坐标体系。\n")

    # ── 断点检测 ──
    has_checkpoint = STEP1_FILE.exists() or STEP2_FILE.exists()
    step3_only = False
    if has_checkpoint:
        existing = [f.name for f in get_checkpoint_files() if f.exists()]

        if ask_checkpoint_fn is not None:
            # GUI 模式：弹窗选择
            choice = ask_checkpoint_fn(existing)
        else:
            # CLI 模式：文本菜单
            log_fn(f"发现上次未完成的进度（{', '.join(existing)}）\n")
            log_fn("  c  继续上次进度（自动跳过已完成步骤）")
            log_fn("")
            log_fn("  ── 或从指定步骤重新开始 ──")
            log_fn("  1  Step 1：重新翻页搜索贝壳小区列表，获取所有小区 URL")
            log_fn("             （清除全部断点，联网爬取，耗时最长）")
            log_fn("  2  Step 2：重新逐一访问小区详情页")
            log_fn("             获取楼栋数 / 总户数 / 坐标等原始数据，保存全部小区")
            log_fn("             （保留 Step 1 URL 列表，需要联网，耗时较长）")
            log_fn("  3  Step 3：重新计算距离 + 筛选 + 生成 Excel")
            log_fn(f"             计算各小区与目标距离，按当前半径 {config.RADIUS_KM}km 筛选，")
            log_fn("             计算人口估算并导出 Excel")
            log_fn("             （不联网，秒级完成；改了半径 / 坐标 / 公式后用此选项）")
            log_fn("")
            log_fn("  r  完全重置（同 1，清除全部断点从头开始）")
            log_fn("  q  退出")
            log_fn("")
            print("请选择 [c/1/2/3/r/q]: ", end="", flush=True)
            while True:
                choice = input().strip().lower()
                if choice in ("c", "1", "2", "3", "r", "q"):
                    break
                print("请输入 c / 1 / 2 / 3 / r / q: ", end="", flush=True)

        # 根据选择处理断点文件
        if choice == "c":
            pass
        elif choice in ("1", "r"):
            for f in get_checkpoint_files():
                if f.exists():
                    f.unlink()
            log_fn("已清除所有断点，从 Step 1 开始。\n")
        elif choice == "2":
            if STEP2_FILE.exists():
                STEP2_FILE.unlink()
            log_fn("已清除 Step 2 断点，将重新抓取小区详情（保留 Step 1 URL 列表）。\n")
        elif choice == "3":
            if not STEP2_FILE.exists():
                log_fn("  ⚠ 尚无 Step 2 数据，将先运行 Step 2（抓取详情页）再生成 Excel。\n")
            else:
                log_fn(f"  保留 Step 1/2 数据，按当前半径 {config.RADIUS_KM}km 重新筛选并生成 Excel。\n")
                step3_only = True
        elif choice == "q":
            sys.exit(0)
    else:
        log_fn("当前数据文件夹没有断点文件，将从 Step 1 开始。\n")

    if step3_only:
        data = _load_json(STEP2_FILE)
        run_step3(data["communities"], log_fn=log_fn)
        return

    # 同步浏览器 Cookie
    n = sync_browser_cookies()
    if n > 0:
        log_fn(f"✓ 已从浏览器同步 {n} 个登录 Cookie\n")
    else:
        log_fn("⚠ 未能同步浏览器 Cookie（若尚未在 Chrome 登录贝壳，遇到验证时再处理即可）\n")

    urls = run_step1(log_fn=log_fn, ask_fn=ask_continue_fn)
    communities = run_step2(urls, log_fn=log_fn, ask_fn=ask_continue_fn, max_communities=max_communities)
    run_step3(communities, log_fn=log_fn)


def main(log_fn=print, ask_continue_fn=None, ask_checkpoint_fn=None,
         verify_fn=None, max_communities=None, data_dir=None):
    try:
        return _main_impl(
            log_fn=log_fn,
            ask_continue_fn=ask_continue_fn,
            ask_checkpoint_fn=ask_checkpoint_fn,
            verify_fn=verify_fn,
            max_communities=max_communities,
            data_dir=data_dir,
        )
    finally:
        close_driver()


if __name__ == "__main__":
    cli_flags = {"--cli", "--test", "--reset"}
    if cli_flags & set(sys.argv[1:]):
        main()
    else:
        from gui import ScraperGUI
        app = ScraperGUI()
        app.run(main)
