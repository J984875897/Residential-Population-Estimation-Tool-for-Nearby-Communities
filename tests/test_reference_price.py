import tempfile
import unittest
from pathlib import Path

import openpyxl
from bs4 import BeautifulSoup

import main
import scraper


class ReferencePriceTests(unittest.TestCase):
    def test_extract_reference_price_from_detail_text(self):
        soup = BeautifulSoup(
            """
            <html>
              <body>
                <div>参考均价</div>
                <span>58,321元/㎡</span>
              </body>
            </html>
            """,
            "html.parser",
        )

        self.assertEqual(scraper._extract_reference_price(soup), 58321)

    def test_write_excel_includes_reference_price_column(self):
        rows = [
            {
                "name": "测试小区",
                "distance_km": 0.321,
                "in_range": True,
                "N": 3,
                "a": 456,
                "b": 2.35,
                "b_note": None,
                "P": 1234,
                "reference_price": 58321,
                "url": "https://gz.ke.com/xiaoqu/123/",
            }
        ]

        with tempfile.TemporaryDirectory() as tmp:
            output_file = Path(tmp) / "result.xlsx"
            main._write_excel(rows, 1234, str(output_file))

            ws = openpyxl.load_workbook(output_file).active
            headers = [ws.cell(row=1, column=col).value for col in range(1, 10)]

        self.assertIn("参考均价(元/㎡)", headers)
        price_col = headers.index("参考均价(元/㎡)") + 1
        self.assertEqual(ws.cell(row=2, column=price_col).value, 58321)


if __name__ == "__main__":
    unittest.main()
